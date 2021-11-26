import datetime
from io import BytesIO

from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.worksheet.cell_range import CellRange
from xls2xlsx import XLS2XLSX

from Object.models import *

def form_ks2(request, name, file, file_smeta):
    """
    View filling KS2_KS3 templates
    """

    object = Object.objects.get(name_object=name)

    if file.name.split('.')[1] == 'xlsx':
        loaded_wb = load_workbook(file)
    elif file.name.split('.')[1] == 'xls':
        x2x = XLS2XLSX(file)
        loaded_wb = x2x.to_xlsx()
    else:
        raise ValueError('Uncorrect format of input .xls file!')

    if file_smeta.name.split('.')[1] == 'xlsx':
        loaded_smeta = load_workbook(file_smeta)
    elif file_smeta.name.split('.')[1] == 'xls':
        x2x = XLS2XLSX(file_smeta)
        loaded_smeta = x2x.to_xlsx()
    else:
        raise ValueError('Uncorrect format of input .xls file!')

    wb = load_workbook('KS2_KS3.xlsx')

    """
    KS2
    """

    sheet_ranges = wb['Акт по форме КС-2']
    ws = wb.active
    loaded_sheet_ranges = loaded_wb['Акт по форме КС-2']

    if object.zakazchik:
        sheet_ranges['B7'].value = sheet_ranges['B7'].value.replace(\
            'полеЗаказчик',object.zakazchik)
    else:
        sheet_ranges['B7'].value = sheet_ranges['B7'].value.replace(\
            'полеЗаказчик','')

    if object.kontragent:
        if (object.kontragent.name_kontragent or object.kontragent.INN
            or object.kontragent.KPP or object.kontragent.Ur_address
            or object.kontragent.telephone):
            sheet_ranges['B8'].value = sheet_ranges['B8'].value.replace(
                'полеПодрядчик',f'{object.kontragent.name_kontragent}, '
                + f'ИНН {object.kontragent.INN}/ КПП {object.kontragent.KPP}, '
                + f'адрес: {object.kontragent.Ur_address}, '
                + f'тел:{object.kontragent.telephone}')
        else:
            sheet_ranges['B8'].value = sheet_ranges['B8'].value.replace(
                'полеПодрядчик','')
    else:
        sheet_ranges['B8'].value = sheet_ranges['B8'].value.replace(
            'полеПодрядчик','')

    if object.ks2_stroika:
        sheet_ranges['C9'].value = sheet_ranges['C9'].value.replace('полеСтройка',
                                                                object.ks2_stroika)
    else:
        sheet_ranges['C9'].value = sheet_ranges['C9'].value.replace('полеСтройка','')

    if object.ks2_object:
        sheet_ranges['C10'].value = sheet_ranges['C10'].value.replace(
            'полеОбъект', object.ks2_object)
    else:
        sheet_ranges['C10'].value = sheet_ranges['C10'].value.replace(
            'полеОбъект', '')

    if object.smeta:
        if object.smeta.nomer_dogovor:
            sheet_ranges['F12'].value = sheet_ranges['F12'].value.replace(
                'полеНомердоговораподр', object.smeta.nomer_dogovor)
        else:
            sheet_ranges['F12'].value = sheet_ranges['F12'].value.replace(
                'полеНомердоговораподр', '')
    else:
        sheet_ranges['F12'].value = sheet_ranges['F12'].value.replace(
            'полеНомердоговораподр', '')

    if object.smeta:
        if object.smeta.date_dogovor:
            sheet_ranges['F13'].value = sheet_ranges['F13'].value.replace(
                'полеДатадоговораподр',datetime.datetime.strftime(
                    object.smeta.date_dogovor,"%d.%m.%Y"))
        else:
            sheet_ranges['F13'].value = sheet_ranges['F13'].value.replace(
                'полеДатадоговораподр', '')
    else:
        sheet_ranges['F13'].value = sheet_ranges['F13'].value.replace(
            'полеДатадоговораподр', '')

    if object.smeta:
        if object.smeta.date_nach_zakr:
            sheet_ranges['G18'].value = sheet_ranges['G18'].value.replace(
                    'полеДатазам1',datetime.datetime.strftime(
                    object.smeta.date_nach_zakr,"%d.%m.%Y"))
        else:
            sheet_ranges['G18'].value = sheet_ranges['G18'].value.replace(
                'полеДатазам1', '')
    else:
        sheet_ranges['G18'].value = sheet_ranges['G18'].value.replace(
            'полеДатазам1', '')

    if object.smeta:
        if object.smeta.date_kon_zakr:
            sheet_ranges['H18'].value = sheet_ranges['H18'].value.replace(
                'полеДатазам2',datetime.datetime.strftime(
                object.smeta.date_kon_zakr,"%d.%m.%Y"))
        else:
            sheet_ranges['H18'].value = sheet_ranges['H18'].value.replace(
                'полеДатазам2', '')
    else:
        sheet_ranges['H18'].value = sheet_ranges['H18'].value.replace(
            'полеДатазам2', '')

    start_index3 = 10
    end_index_template = 0
    for start_index4 in range(start_index3, 250):

        if end_index_template:
            break

        if loaded_sheet_ranges[f"A{start_index4}"].value:
            if end_index_template:
                break

            if 'Номер' in str(loaded_sheet_ranges[f"A{start_index4}"].value):
                if end_index_template:
                    break

                start_index_template = 28

                for start_index5 in range(start_index4+4, 250):
                    if loaded_sheet_ranges[f"A{start_index5}"].value:

                        if str(loaded_sheet_ranges[f"A{start_index5}"].value).isdigit():
                            for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'):
                                sheet_ranges[f'{letter}{start_index_template}'].value = loaded_sheet_ranges[f"{letter}{start_index5}"].value
                        else:

                            sheet_ranges[f'A{start_index_template}'].value = loaded_sheet_ranges[f"A{start_index5}"].value
                            sheet_ranges[f'A{start_index_template}'].alignment = Alignment(horizontal='left')

                            if loaded_sheet_ranges[f"H{start_index5}"].value:
                                sheet_ranges.merge_cells(f'A{start_index_template}:G{start_index_template}')
                                sheet_ranges[f'H{start_index_template}'].value = loaded_sheet_ranges[f"H{start_index5}"].value
                                sheet_ranges[f'H{start_index_template}'].alignment = Alignment(horizontal='right')
                                if loaded_sheet_ranges[f"A{start_index5}"].font.bold:
                                    sheet_ranges[f'A{start_index_template}'].font = Font(name='Arial', size=10, bold=True)
                                    sheet_ranges[f'H{start_index_template}'].font = Font(name='Arial', size=9, bold=True)
                                else:
                                    sheet_ranges[f'A{start_index_template}'].font = Font(name='Arial', size=10, bold=False)
                                    sheet_ranges[f'H{start_index_template}'].font = Font(name='Arial', size=9, bold=False)
                                if 'Раздел' in str(loaded_sheet_ranges[f"A{start_index5}"].value):
                                    sheet_ranges[f"A{start_index_template}"].font = Font(name='Arial', size=11, bold=True)

                            else:
                                sheet_ranges.merge_cells(f'A{start_index_template}:H{start_index_template}')
                                if loaded_sheet_ranges[f"A{start_index5}"].font.bold:
                                    sheet_ranges[f'A{start_index_template}'].font = Font(name='Arial', size=10, bold=True)
                                else:
                                    sheet_ranges[f'A{start_index_template}'].font = Font(name='Arial', size=10, bold=False)
                                if 'Раздел' in str(loaded_sheet_ranges[f"A{start_index5}"].value):
                                    sheet_ranges[f"A{start_index_template}"].font = Font(name='Arial', size=11, bold=True)

                        start_index_template += 1

                    else:
                        end_index_template = start_index_template
                        break
            else:
                if end_index_template:
                    break
        else:
            if end_index_template:
                break

    black_side = Side(border_style="thin", color="000000")

    sheet_ranges.delete_rows(end_index_template, 191-end_index_template)

    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'):
        sheet_ranges[f'{letter}{end_index_template-1}'].border = Border(top=black_side,
            left=black_side, right=black_side, bottom=black_side)

    if object.smeta:
        if object.smeta.date_ks2:
            sheet_ranges['F18'].value = sheet_ranges['F18'].value.replace(
                'полеДатаКС2',datetime.datetime.strftime(object.smeta.date_ks2,"%d.%m.%Y"))
        else:
            sheet_ranges['F18'].value = sheet_ranges['F18'].value.replace(
                'полеДатаКС2', '')
    else:
        sheet_ranges['F18'].value = sheet_ranges['F18'].value.replace(
            'полеДатаКС2', '')

    count = 1
    start_value = ''
    start_position = ''
    while (start_value == ''):
        if count == 150:
            break
        if sheet_ranges[f'A{count}'].value:

            if 'Сдал:' in str(sheet_ranges[f'A{count}'].value):
                start_value = 'Сдал:'
                start_position = f'A{count}'
                if object.ks2_podryadchik:
                    if object.ks2_podryadchik.post:
                        sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                            'полеКС2подрядчик', object.ks2_podryadchik.post)

                        sheet_ranges.merge_cells(f'A{count}:C{count+2}')
                        sheet_ranges.merge_cells(f'D{count}:E{count}')
                        sheet_ranges.merge_cells(f'D{count+1}:E{count+1}')
                    else:
                        sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                        'полеКС2подрядчик', '')
                else:
                    sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                    'полеКС2подрядчик', '')

                if object.ks2_podryadchik:
                    if object.ks2_podryadchik.fio:
                        sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                                    'полеКС2подрядчикФИО', object.ks2_podryadchik.fio)
                    else:
                        sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                                    'полеКС2подрядчикФИО', '')
                else:
                    sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                                'полеКС2подрядчикФИО', '')
            else:
                count += 1
        else:
            count += 1

    while (start_value == 'Сдал:'):
        if sheet_ranges[f'A{count}'].value:
            if 'Принял:' in str(sheet_ranges[f'A{count}'].value):
                start_value = 'Принял:'
                start_position = f'A{count}'
                if object.ks2_zakazchik:
                    if object.ks2_zakazchik.post:
                        sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                            'полеКС2заказчик',object.ks2_zakazchik.post)

                        sheet_ranges.merge_cells(f'A{count}:C{count+2}')
                        sheet_ranges.merge_cells(f'D{count}:E{count}')
                        sheet_ranges.merge_cells(f'D{count+1}:E{count+1}')
                    else:
                        sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                            'полеКС2заказчик', '')
                else:
                    sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                        'полеКС2заказчик', '')

                if object.ks2_zakazchik:
                    if object.ks2_zakazchik.fio:
                        sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                            'полеКС2заказчикФИО',object.ks2_zakazchik.fio)
                    else:
                        sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                            'полеКС2заказчикФИО', '')
                else:
                    sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                        'полеКС2заказчикФИО', '')

            else:
                count += 1
        else:
            count += 1

    while (start_value == 'Принял:'):
        if sheet_ranges[f'A{count}'].value:
            if 'полеТехнадзор' in str(sheet_ranges[f'A{count}'].value):
                start_value = 'полеТехнадзор'
                start_position = f'A{count}'
                if object.tehnadzor:
                    if object.tehnadzor.person:
                        if object.tehnadzor.person.post:
                            sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                                'полеТехнадзор',object.tehnadzor.person.post)

                            sheet_ranges.merge_cells(f'A{count}:C{count+2}')
                            sheet_ranges.merge_cells(f'D{count}:E{count}')
                            sheet_ranges.merge_cells(f'D{count+1}:E{count+1}')
                        else:
                            sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                                'полеТехнадзор','')
                    else:
                        sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                            'полеТехнадзор','')
                else:
                    sheet_ranges[f'A{count}'].value = sheet_ranges[f'A{count}'].value.replace(
                        'полеТехнадзор','')

                if object.tehnadzor:
                    if object.tehnadzor.person:
                        if object.tehnadzor.person.fio:
                            sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                                'полеТехнадзорФИО',object.tehnadzor.person.fio)
                        else:
                            sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                                'полеТехнадзорФИО', '')
                    else:
                        sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                            'полеТехнадзорФИО', '')
                else:
                    sheet_ranges[f'F{count}'].value = sheet_ranges[f'F{count}'].value.replace(
                        'полеТехнадзорФИО', '')
            else:
                count += 1
        else:
            count += 1

    """
    KS3
    """

    sheet_ranges_ks3 = wb['КС3']

    if object.zakazchik:
        sheet_ranges_ks3['A9'].value = sheet_ranges_ks3['A9'].value.replace(
            'полеЗаказчик',object.zakazchik)
    else:
        sheet_ranges_ks3['A9'].value = sheet_ranges_ks3['A9'].value.replace(
            'полеЗаказчик','')

    if object.kontragent:
        if (object.kontragent.name_kontragent or object.kontragent.INN
            or object.kontragent.KPP or object.kontragent.Ur_address
            or object.kontragent.telephone):
            sheet_ranges_ks3['A11'].value = sheet_ranges_ks3['A11'].value.replace(
                'полеПодрядчик',f'{object.kontragent.name_kontragent}, '
                + f'ИНН {object.kontragent.INN}/ КПП {object.kontragent.KPP}, '
                + f'адрес: {object.kontragent.Ur_address}, '
                + f'тел:{object.kontragent.telephone}')
        else:
            sheet_ranges_ks3['A11'].value = sheet_ranges_ks3['A11'].value.replace(
                'полеПодрядчик', '')
    else:
        sheet_ranges_ks3['A11'].value = sheet_ranges_ks3['A11'].value.replace(
            'полеПодрядчик', '')

    if object.ks2_stroika:
        sheet_ranges_ks3['A13'].value = sheet_ranges_ks3['A13'].value.replace(
            'полеСтройка',object.ks2_stroika)
    else:
        sheet_ranges_ks3['A13'].value = sheet_ranges_ks3['A13'].value.replace(
            'полеСтройка', '')

    if object.ks2_object:
        sheet_ranges_ks3['A15'].value = sheet_ranges_ks3['A15'].value.replace(
            'полеОбъект',object.ks2_object)
    else:
        sheet_ranges_ks3['A15'].value = sheet_ranges_ks3['A15'].value.replace(
            'полеОбъект', '')

    if object.smeta:
        if object.smeta.nomer_dogovor:
            sheet_ranges_ks3['I17'].value = sheet_ranges_ks3['I17'].value.replace(
                'полеНомердоговораподр', object.smeta.nomer_dogovor)
        else:
            sheet_ranges_ks3['I17'].value = sheet_ranges_ks3['I17'].value.replace(
                'полеНомердоговораподр', '')
    else:
        sheet_ranges_ks3['I17'].value = sheet_ranges_ks3['I17'].value.replace(
            'полеНомердоговораподр', '')

    if object.smeta:
        if object.smeta.date_dogovor:
            sheet_ranges_ks3['I18'].value = sheet_ranges_ks3['I18'].value.replace(
                    'дд',datetime.datetime.strftime(object.smeta.date_dogovor,"%d"))
            sheet_ranges_ks3['J18'].value = sheet_ranges_ks3['J18'].value.replace(
                    'мм',datetime.datetime.strftime(object.smeta.date_dogovor,"%m"))
            sheet_ranges_ks3['K18'].value = sheet_ranges_ks3['K18'].value.replace(
                    'гггг',datetime.datetime.strftime(object.smeta.date_dogovor,"%Y"))
        else:
            sheet_ranges_ks3['I18'].value = sheet_ranges_ks3['I18'].value.replace(
                'дд', '')
            sheet_ranges_ks3['J18'].value = sheet_ranges_ks3['J18'].value.replace(
                'мм', '')
            sheet_ranges_ks3['K18'].value = sheet_ranges_ks3['K18'].value.replace(
                'гггг', '')
    else:
        sheet_ranges_ks3['I18'].value = sheet_ranges_ks3['I18'].value.replace(
            'дд', '')
        sheet_ranges_ks3['J18'].value = sheet_ranges_ks3['J18'].value.replace(
            'мм', '')
        sheet_ranges_ks3['K18'].value = sheet_ranges_ks3['K18'].value.replace(
            'гггг', '')

    if object.ks2_podryadchik:
        if object.ks2_podryadchik.post:
            sheet_ranges_ks3['A40'].value = sheet_ranges_ks3['A40'].value.replace(
                'полеКС2подрядчик',object.ks2_podryadchik.post)
        else:
            sheet_ranges_ks3['A40'].value = sheet_ranges_ks3['A40'].value.replace(
                'полеКС2подрядчик', '')

        if object.ks2_podryadchik.fio:
            sheet_ranges_ks3['H40'].value = sheet_ranges_ks3['H40'].value.replace(
                'полеКС2подрядчикФИО',object.ks2_podryadchik.fio)
        else:
            sheet_ranges_ks3['H40'].value = sheet_ranges_ks3['H40'].value.replace(
                'полеКС2подрядчикФИО', '')
    else:
        sheet_ranges_ks3['A40'].value = sheet_ranges_ks3['A40'].value.replace(
            'полеКС2подрядчик', '')
        sheet_ranges_ks3['H40'].value = sheet_ranges_ks3['H40'].value.replace(
            'полеКС2подрядчикФИО', '')

    if object.ks2_zakazchik:
        if object.ks2_zakazchik.post:
            sheet_ranges_ks3['A46'].value = sheet_ranges_ks3['A46'].value.replace(
                'полеКС2заказчик',object.ks2_zakazchik.post)
        else:
            sheet_ranges_ks3['A46'].value = sheet_ranges_ks3['A46'].value.replace(
                'полеКС2заказчик', '')

        if object.ks2_zakazchik.fio:
            sheet_ranges_ks3['H46'].value = sheet_ranges_ks3['H46'].value.replace(
                'полеКС2заказчикФИО',object.ks2_zakazchik.fio)
        else:
            sheet_ranges_ks3['H46'].value = sheet_ranges_ks3['H46'].value.replace(
                'полеКС2заказчикФИО', '')
    else:
        sheet_ranges_ks3['A46'].value = sheet_ranges_ks3['A46'].value.replace(
            'полеКС2заказчик', '')
        sheet_ranges_ks3['H46'].value = sheet_ranges_ks3['H46'].value.replace(
            'полеКС2заказчикФИО', '')

    if object.smeta:
        if object.smeta.date_ks2:
            sheet_ranges_ks3['F23'].value = sheet_ranges_ks3['F23'].value.replace(
                'полеДатаКС2',datetime.datetime.strftime(object.smeta.date_ks2,"%d.%m.%Y"))
        else:
            sheet_ranges_ks3['F23'].value = sheet_ranges_ks3['F23'].value.replace(
                'полеДатаКС2', '')
    else:
        sheet_ranges_ks3['F23'].value = sheet_ranges_ks3['F23'].value.replace(
            'полеДатаКС2', '')

    if object.smeta:
        if object.smeta.date_nach_zakr:
            sheet_ranges_ks3['H23'].value = sheet_ranges_ks3['H23'].value.replace(
                'полеДатазам1',datetime.datetime.strftime(object.smeta.date_nach_zakr,"%d.%m.%Y"))
        else:
            sheet_ranges_ks3['H23'].value = sheet_ranges_ks3['H23'].value.replace(
                'полеДатазам1','')
    else:
        sheet_ranges_ks3['H23'].value = sheet_ranges_ks3['H23'].value.replace(
            'полеДатазам1','')

    if object.smeta:
        if object.smeta.date_kon_zakr:
            sheet_ranges_ks3['I23'].value = sheet_ranges_ks3['I23'].value.replace(
                'полеДатазам2',datetime.datetime.strftime(object.smeta.date_kon_zakr,"%d.%m.%Y"))
        else:
            sheet_ranges_ks3['I23'].value = sheet_ranges_ks3['I23'].value.replace(
                'полеДатазам2','')
    else:
        sheet_ranges_ks3['I23'].value = sheet_ranges_ks3['I23'].value.replace(
            'полеДатазам2','')

    if object.ks2_object:
        sheet_ranges_ks3['B31'].value = sheet_ranges_ks3['B31'].value.replace(
            'полеОбъект',object.ks2_object)
        sheet_ranges_ks3['B31'].alignment = Alignment(wrapText=True)
    else:
        sheet_ranges_ks3['B31'].value = sheet_ranges_ks3['B31'].value.replace(
            'полеОбъект', '')

    if object.smeta:
        if object.smeta.summa_ks2_bez_nds:
            sheet_ranges_ks3['F30'].value = sheet_ranges_ks3['F30'].value.replace(
                'полеСуммутвсметы', str(object.smeta.summa_ks2_bez_nds).replace('.',','))
            sheet_ranges_ks3['H30'].value = sheet_ranges_ks3['H30'].value.replace(
                'полеСуммутвсметы', str(object.smeta.summa_ks2_bez_nds).replace('.',','))
            sheet_ranges_ks3['I30'].value = sheet_ranges_ks3['I30'].value.replace(
                'полеСуммутвсметы', str(object.smeta.summa_ks2_bez_nds).replace('.',','))
            sheet_ranges_ks3['F31'].value = sheet_ranges_ks3['F31'].value.replace(
                'полеСуммутвсметы', str(object.smeta.summa_ks2_bez_nds).replace('.',','))
            sheet_ranges_ks3['H31'].value = sheet_ranges_ks3['H31'].value.replace(
                'полеСуммутвсметы', str(object.smeta.summa_ks2_bez_nds).replace('.',','))
            sheet_ranges_ks3['I31'].value = sheet_ranges_ks3['I31'].value.replace(
                'полеСуммутвсметы', str(object.smeta.summa_ks2_bez_nds).replace('.',','))
        else:
            sheet_ranges_ks3['F30'].value = sheet_ranges_ks3['F30'].value.replace(
                'полеСуммутвсметы', '')
            sheet_ranges_ks3['H30'].value = sheet_ranges_ks3['H30'].value.replace(
                'полеСуммутвсметы', '')
            sheet_ranges_ks3['I30'].value = sheet_ranges_ks3['I30'].value.replace(
                'полеСуммутвсметы', '')
            sheet_ranges_ks3['F31'].value = sheet_ranges_ks3['F31'].value.replace(
                'полеСуммутвсметы', '')
            sheet_ranges_ks3['H31'].value = sheet_ranges_ks3['H31'].value.replace(
                'полеСуммутвсметы', '')
            sheet_ranges_ks3['I31'].value = sheet_ranges_ks3['I31'].value.replace(
                'полеСуммутвсметы', '')

        if object.smeta.summa_ks2_bez_nds:
            a =  str(object.smeta.summa_ks2_bez_nds).replace('.',',')
            sheet_ranges_ks3['I32'].value = sheet_ranges_ks3['I32'].value.replace(
                'полеКС2безндс', a)
            b = float(object.smeta.summa_ks2_bez_nds.replace(',','.'))
            b2 = round(b*0.2,2)
            d = str(b2)
            d = d.replace('.',',')
            sheet_ranges_ks3['I33'].value = sheet_ranges_ks3['I33'].value.replace(
                'полеСуммаНДС', d)
            c1 = float(object.smeta.summa_ks2_bez_nds.replace(',','.'))
            c = str(round(c1*0.2 + c1, 2)).replace('.',',')
            sheet_ranges_ks3['I34'].value = sheet_ranges_ks3['I34'].value.replace(
                'полеВсегосучетомНДС', c)

        else:
            sheet_ranges_ks3['I32'].value = sheet_ranges_ks3['I32'].value.replace(
                'полеКС2безндс', '')
            sheet_ranges_ks3['I33'].value = sheet_ranges_ks3['I33'].value.replace(
                'полеСуммаНДС', '')
            sheet_ranges_ks3['I34'].value = sheet_ranges_ks3['I34'].value.replace(
                'полеВсегосучетомНДС', '')
    else:
        sheet_ranges_ks3['F30'].value = sheet_ranges_ks3['F30'].value.replace(
            'полеСуммутвсметы', '')
        sheet_ranges_ks3['H30'].value = sheet_ranges_ks3['H30'].value.replace(
            'полеСуммутвсметы', '')
        sheet_ranges_ks3['I30'].value = sheet_ranges_ks3['I30'].value.replace(
            'полеСуммутвсметы', '')
        sheet_ranges_ks3['F31'].value = sheet_ranges_ks3['F31'].value.replace(
            'полеСуммутвсметы', '')
        sheet_ranges_ks3['H31'].value = sheet_ranges_ks3['H31'].value.replace(
            'полеСуммутвсметы', '')
        sheet_ranges_ks3['I31'].value = sheet_ranges_ks3['I31'].value.replace(
            'полеСуммутвсметы', '')
        sheet_ranges_ks3['I32'].value = sheet_ranges_ks3['I32'].value.replace(
            'полеКС2безндс', '')
        sheet_ranges_ks3['I33'].value = sheet_ranges_ks3['I33'].value.replace(
            'полеСуммаНДС', '')
        sheet_ranges_ks3['I34'].value = sheet_ranges_ks3['I34'].value.replace(
            'полеВсегосучетомНДС', '')
    
    """
    СМЕТА
    """
    
    sheet_ranges_smeta = wb['ЛСР 13 граф']
    loaded_sheet_ranges_smeta = loaded_smeta.active

    if object.ks2_podryadchik:
        if object.ks2_podryadchik.post:
            sheet_ranges_smeta['A2'].value = sheet_ranges_smeta['A2'].value.replace(
                'полеПодрядчикдол',object.ks2_podryadchik.post)
        else:
            sheet_ranges_smeta['A2'].value = sheet_ranges_smeta['A2'].value.replace(
                'полеПодрядчикдол', '')

        if object.ks2_podryadchik.fio:
            sheet_ranges_smeta['C4'].value = sheet_ranges_smeta['C4'].value.replace(
                'полеПодрядчикФИО',object.ks2_podryadchik.fio)
        else:
            sheet_ranges_smeta['C4'].value = sheet_ranges_smeta['C4'].value.replace(
                'полеПодрядчикФИО', '')
    else:
        sheet_ranges_smeta['A2'].value = sheet_ranges_smeta['A2'].value.replace(
            'полеПодрядчикдол', '')
        sheet_ranges_smeta['C4'].value = sheet_ranges_smeta['C4'].value.replace(
            'полеПодрядчикФИО', '')

    if object.ks2_zakazchik:
        if object.ks2_zakazchik.post:
            sheet_ranges_smeta['G2'].value = sheet_ranges_smeta['G2'].value.replace(
                'полеЗаказчикдол',object.ks2_zakazchik.post)
        else:
            sheet_ranges_smeta['G2'].value = sheet_ranges_smeta['G2'].value.replace(
                'полеЗаказчикдол', '')

        if object.ks2_zakazchik.fio:
            sheet_ranges_smeta['K4'].value = sheet_ranges_smeta['K4'].value.replace(
                'полеЗаказчикФИО',object.ks2_zakazchik.fio)
        else:
            sheet_ranges_smeta['K4'].value = sheet_ranges_smeta['K4'].value.replace(
                'полеЗаказчикФИО', '')
    else:
        sheet_ranges_smeta['G2'].value = sheet_ranges_smeta['G2'].value.replace(
            'полеЗаказчикдол', '')
        sheet_ranges_smeta['K4'].value = sheet_ranges_smeta['K4'].value.replace(
            'полеЗаказчикФИО', '')

    sheet_ranges_smeta['A5'].value = sheet_ranges_smeta['A5'].value.replace(
        'ГОД', str(datetime.datetime.today().year))

    sheet_ranges_smeta['I5'].value = sheet_ranges_smeta['I5'].value.replace(
        'ГОД', str(datetime.datetime.today().year))

    if object.ks2_stroika:
        sheet_ranges_smeta['A6'].value = sheet_ranges_smeta['A6'].value.replace(
            'полеНазваниеобъекта', object.ks2_stroika)
    else:
        sheet_ranges_smeta['A6'].value = sheet_ranges_smeta['A6'].value.replace(
            'полеНазваниеобъекта', '')

    if object.place:
        sheet_ranges_smeta['C12'].value = sheet_ranges_smeta['C12'].value.replace(
            'полеМестоположениеобъекта', object.place)
    else:
        sheet_ranges_smeta['C12'].value = sheet_ranges_smeta['C12'].value.replace(
            'полеМестоположениеобъекта', '')

    for start_index_smeta in range(1,50):
        if loaded_sheet_ranges_smeta[f"C{start_index_smeta}"].value:
            if 'Сметная стоимость' in str(loaded_sheet_ranges_smeta[f"C{start_index_smeta}"].value):
                sheet_ranges_smeta['E16'].value = sheet_ranges_smeta['E16'].value.replace(
                    'полеЦена1', loaded_sheet_ranges_smeta[f"E{start_index_smeta}"].value)
                sheet_ranges_smeta['E17'].value = sheet_ranges_smeta['E17'].value.replace(
                    'полеЦена2', loaded_sheet_ranges_smeta[f"E{start_index_smeta+1}"].value)
                sheet_ranges_smeta['E18'].value = sheet_ranges_smeta['E18'].value.replace(
                    'полеЦена3', loaded_sheet_ranges_smeta[f"E{start_index_smeta+2}"].value)
                sheet_ranges_smeta['E19'].value = sheet_ranges_smeta['E19'].value.replace(
                    'полеЦена4', loaded_sheet_ranges_smeta[f"E{start_index_smeta+3}"].value)
                sheet_ranges_smeta['E20'].value = sheet_ranges_smeta['E20'].value.replace(
                    'полеЦена5', loaded_sheet_ranges_smeta[f"E{start_index_smeta+4}"].value)
                break
        else:
            continue

    if 'полеЦена1' in str(sheet_ranges_smeta['E16'].value):
        sheet_ranges_smeta['E16'].value = sheet_ranges_smeta['E16'].value.replace(
            'полеЦена1', '')
    if 'полеЦена2' in str(sheet_ranges_smeta['E17'].value):
        sheet_ranges_smeta['E17'].value = sheet_ranges_smeta['E17'].value.replace(
            'полеЦена2', '')
    if 'полеЦена3' in str(sheet_ranges_smeta['E18'].value):
        sheet_ranges_smeta['E18'].value = sheet_ranges_smeta['E18'].value.replace(
            'полеЦена3', '')
    if 'полеЦена4' in str(sheet_ranges_smeta['E19'].value):
        sheet_ranges_smeta['E19'].value = sheet_ranges_smeta['E19'].value.replace(
            'полеЦена4', '')
    if 'полеЦена5' in str(sheet_ranges_smeta['E20'].value):
        sheet_ranges_smeta['E20'].value = sheet_ranges_smeta['E20'].value.replace(
            'полеЦена5', '')

    for start_index_smeta2 in range(start_index_smeta,50):
        if loaded_sheet_ranges_smeta[f"A{start_index_smeta2}"].value:
            if '№ пп' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta2}"].value):
                break
            else:
                continue
        else:
            continue

    start_index_table_templ = 28

    black_side = Side(border_style="thin", color="000000")

    for start_index_smeta3 in range(start_index_smeta2+4, 300):

        if CellRange(f'A{start_index_table_templ}:H{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:H{start_index_table_templ}')
        if CellRange(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}')

        if CellRange(f'A{start_index_table_templ}:G{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:G{start_index_table_templ}')
        if CellRange(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}')

        if loaded_sheet_ranges_smeta[f"J{start_index_smeta3}"].value:

            sheet_ranges_smeta[f'J{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"J{start_index_smeta3}"].value

            if 'Итого по разделу' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value)\
                or 'ВСЕГО по смете' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value):

                sheet_ranges_smeta[f'J{start_index_table_templ}'].font = Font(name='Arial', size=8, bold=True)
            else:
                sheet_ranges_smeta[f'J{start_index_table_templ}'].font = Font(name='Arial', size=8, bold=False)

            sheet_ranges_smeta[f'J{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)

        if loaded_sheet_ranges_smeta[f"K{start_index_smeta3}"].value:
            sheet_ranges_smeta[f'K{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"K{start_index_smeta3}"].value
            sheet_ranges_smeta[f'K{start_index_table_templ}'].font = Font(name='Arial', size=8)
            sheet_ranges_smeta[f'K{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)

        if loaded_sheet_ranges_smeta[f"L{start_index_smeta3}"].value:
            sheet_ranges_smeta[f'L{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"L{start_index_smeta3}"].value
            sheet_ranges_smeta[f'L{start_index_table_templ}'].font = Font(name='Arial', size=8)
            sheet_ranges_smeta[f'L{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)

        if loaded_sheet_ranges_smeta[f"M{start_index_smeta3}"].value:
            sheet_ranges_smeta[f'M{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"M{start_index_smeta3}"].value
            sheet_ranges_smeta[f'M{start_index_table_templ}'].font = Font(name='Arial', size=8)
            sheet_ranges_smeta[f'M{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)

        if loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value:
            if str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value).isdigit():

                sheet_ranges_smeta[f'A{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value
                sheet_ranges_smeta[f'D{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"D{start_index_smeta3}"].value
                sheet_ranges_smeta[f'D{start_index_table_templ}'].font = Font(name='Arial', size=9)
                sheet_ranges_smeta[f'D{start_index_table_templ}'].alignment = Alignment(horizontal='center', vertical='top', wrapText=True)
                sheet_ranges_smeta[f'G{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"G{start_index_smeta3}"].value
                sheet_ranges_smeta[f'G{start_index_table_templ}'].font = Font(name='Arial', size=8)
                sheet_ranges_smeta[f'G{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)
                sheet_ranges_smeta[f'H{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"H{start_index_smeta3}"].value
                sheet_ranges_smeta[f'H{start_index_table_templ}'].font = Font(name='Arial', size=8)
                sheet_ranges_smeta[f'H{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)
                sheet_ranges_smeta[f'I{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"I{start_index_smeta3}"].value
                sheet_ranges_smeta[f'I{start_index_table_templ}'].font = Font(name='Arial', size=8)
                sheet_ranges_smeta[f'I{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)

                slices_B = (str(loaded_sheet_ranges_smeta[f"B{start_index_smeta3}"].value).split('\n'))
                if len(slices_B) > 1:

                    for index_B, slice_B in enumerate(slices_B):

                        if CellRange(f'A{start_index_table_templ}:H{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:H{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}')

                        if CellRange(f'A{start_index_table_templ}:G{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:G{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}')


                        sheet_ranges_smeta[f'B{start_index_table_templ+index_B}'].value = slice_B

                        if index_B == 0:

                            sheet_ranges_smeta[f'B{start_index_table_templ+index_B}'].font = Font(name='Arial', size=9, bold=True)
                            sheet_ranges_smeta[f'B{start_index_table_templ+index_B}'].alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
                        else:

                            sheet_ranges_smeta[f'B{start_index_table_templ+index_B}'].font = Font(name='Arial', size=7, bold=False, italic=True)
                            sheet_ranges_smeta[f'B{start_index_table_templ+index_B}'].alignment = Alignment(horizontal='left',vertical='top', wrapText=True)

                else:

                    sheet_ranges_smeta[f'B{start_index_table_templ}'].value = slices_B[0]
                    sheet_ranges_smeta[f'B{start_index_table_templ}'].font = Font(name='Arial', size=9, bold=True)
                    sheet_ranges_smeta[f'B{start_index_table_templ}'].alignment = Alignment(horizontal='left', wrapText=True)

                slices_C = (str(loaded_sheet_ranges_smeta[f"C{start_index_smeta3}"].value).split('\n'))

                if len(slices_C) != 1:
                    for index_C, slice_C in enumerate(slices_C):

                        if CellRange(f'A{start_index_table_templ}:H{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:H{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}')

                        if CellRange(f'A{start_index_table_templ}:G{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:G{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}')


                        if index_C == 0:
                            sheet_ranges_smeta[f'C{start_index_table_templ+index_C}'].value = slice_C
                            sheet_ranges_smeta[f'C{start_index_table_templ+index_C}'].font = Font(name='Arial', size=9, bold=False)
                            sheet_ranges_smeta[f'C{start_index_table_templ+index_C}'].alignment = Alignment(horizontal='left',vertical='top', wrapText=True)

                        else:
                            string = ''
                            for index, slice in enumerate(slices_C):
                                if index != 0:
                                    string += slice
                                    string += '\n'

                            sheet_ranges_smeta[f'C{start_index_table_templ+index_C}'].value = string
                            sheet_ranges_smeta[f'C{start_index_table_templ+index_C}'].font = Font(name='Arial', size=7, bold=False, italic=True)
                            sheet_ranges_smeta[f'C{start_index_table_templ+index_C}'].alignment = Alignment(horizontal='left',vertical='top', wrapText=True)
                            break

                else:

                    sheet_ranges_smeta[f'C{start_index_table_templ}'].value = slices_C[0]
                    sheet_ranges_smeta[f'C{start_index_table_templ}'].font = Font(name='Arial', size=9, bold=False)
                    sheet_ranges_smeta[f'C{start_index_table_templ}'].alignment = Alignment(horizontal='left', wrapText=True)

                slices_E = (str(loaded_sheet_ranges_smeta[f"E{start_index_smeta3}"].value).split('\n'))

                if len(slices_E) != 1:

                    for index_E, slice_E in enumerate(slices_E):

                        if CellRange(f'A{start_index_table_templ}:H{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:H{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}')

                        if CellRange(f'A{start_index_table_templ}:G{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:G{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}')


                        if index_E == 0:
                            sheet_ranges_smeta[f'E{start_index_table_templ+index_E}'].value = slice_E
                            sheet_ranges_smeta[f'E{start_index_table_templ+index_E}'].font = Font(name='Arial', size=8, bold=False)
                            sheet_ranges_smeta[f'E{start_index_table_templ+index_E}'].alignment = Alignment(horizontal='center',vertical='top', wrapText=True)
                        else:
                            string = ''
                            for index, slice in enumerate(slices_E):
                                if index != 0:
                                    string += slice
                                    string += '\n'

                            sheet_ranges_smeta[f'E{start_index_table_templ+index_E}'].value = string
                            sheet_ranges_smeta[f'E{start_index_table_templ+index_E}'].font = Font(name='Arial', size=6, bold=False, italic=True)
                            sheet_ranges_smeta[f'E{start_index_table_templ+index_E}'].alignment = Alignment(horizontal='center',vertical='top', wrapText=True)
                            break

                else:

                    sheet_ranges_smeta[f'E{start_index_table_templ}'].value = slices_E[0]
                    sheet_ranges_smeta[f'E{start_index_table_templ}'].font = Font(name='Arial', size=8, bold=False)
                    sheet_ranges_smeta[f'E{start_index_table_templ}'].alignment = Alignment(horizontal='center', vertical='top',wrapText=True)

                slices_F = (str(loaded_sheet_ranges_smeta[f"F{start_index_smeta3}"].value).split('\n'))

                if len(slices_F) != 1:

                    for index_F, slice_F in enumerate(slices_F):

                        if CellRange(f'A{start_index_table_templ}:H{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:H{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:H{start_index_table_templ+1}')

                        if CellRange(f'A{start_index_table_templ}:G{start_index_table_templ}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ}:G{start_index_table_templ}')
                        if CellRange(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}') in sheet_ranges_smeta.merged_cells:
                            sheet_ranges_smeta.unmerge_cells(f'A{start_index_table_templ+1}:G{start_index_table_templ+1}')

                        if index_F == 0:
                            sheet_ranges_smeta[f'F{start_index_table_templ+index_F}'].value = slice_F
                            sheet_ranges_smeta[f'F{start_index_table_templ+index_F}'].font = Font(name='Arial', size=8, bold=False)
                            sheet_ranges_smeta[f'F{start_index_table_templ+index_F}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)
                        else:
                            string = ''
                            for index, slice in enumerate(slices_F):
                                if index != 0:
                                    string += slice
                                    string += '\n'

                            sheet_ranges_smeta[f'F{start_index_table_templ+index_F}'].value = string
                            sheet_ranges_smeta[f'F{start_index_table_templ+index_F}'].font = Font(name='Arial', size=6, bold=False, italic=True)
                            sheet_ranges_smeta[f'F{start_index_table_templ+index_F}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)
                            break

                else:

                    sheet_ranges_smeta[f'F{start_index_table_templ}'].value = slices_F[0]
                    sheet_ranges_smeta[f'F{start_index_table_templ}'].font = Font(name='Arial', size=8, bold=False)
                    sheet_ranges_smeta[f'F{start_index_table_templ}'].alignment = Alignment(horizontal='right', vertical='top', wrapText=True)

                for index in range(max([len(slices_B), 2])):
                    if index == 0:

                        sheet_ranges_smeta[f'A{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'B{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'C{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'D{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'E{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'F{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'G{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'H{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'I{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'J{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'K{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'L{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'M{start_index_table_templ+index}'].border = Border(top=black_side,
                            left=black_side, right=black_side)

                    elif index == max([len(slices_B), 2]) - 1:

                        sheet_ranges_smeta[f'A{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'B{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'C{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'D{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'E{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'F{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'G{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'H{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'I{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'J{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'K{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'L{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)
                        sheet_ranges_smeta[f'M{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side, bottom=black_side)

                    else:

                        sheet_ranges_smeta[f'A{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'B{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'C{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'D{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'G{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'H{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'I{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'J{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'K{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'L{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)
                        sheet_ranges_smeta[f'M{start_index_table_templ+index}'].border = Border(
                            left=black_side, right=black_side)

                start_index_table_templ += max([len(slices_B), 2])

            else:

                sheet_ranges_smeta[f'A{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value
                flag = True

                if 'Раздел' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value):
                    sheet_ranges_smeta.merge_cells(f'A{start_index_table_templ}:M{start_index_table_templ}')
                    sheet_ranges_smeta[f"A{start_index_table_templ}"].font = Font(name='Arial', size=10, bold=True)
                    sheet_ranges_smeta[f'A{start_index_table_templ}'].alignment = Alignment(horizontal='left')

                    flag = False

                if 'Итого прямые затраты по разделу' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value)\
                    or 'Итого прямые затраты по смете' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value)\
                    or 'Накладные расходы' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value)\
                    or 'Сметная прибыль' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value)\
                    or 'ВСЕГО по смете' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value)\
                    or 'Итого по разделу' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value):

                    flag = False
                    sheet_ranges_smeta[f'A{start_index_table_templ}'].alignment = Alignment(horizontal='left')

                    if 'Итого по разделу' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value)\
                        or 'ВСЕГО по смете' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value):
                        sheet_ranges_smeta[f'A{start_index_table_templ}'].font = Font(name='Arial', size=9, bold=True)

                    sheet_ranges_smeta.merge_cells(f'A{start_index_table_templ}:I{start_index_table_templ}')

                if 'ИТОГИ ПО СМЕТЕ' in str(loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value):
                    sheet_ranges_smeta[f'A{start_index_table_templ}'].font = Font(name='Arial', size=9, bold=True)
                    sheet_ranges_smeta[f'A{start_index_table_templ}'].alignment = Alignment(horizontal='center')
                    sheet_ranges_smeta.merge_cells(f'A{start_index_table_templ}:M{start_index_table_templ}')
                    flag = False

                if loaded_sheet_ranges_smeta[f"J{start_index_smeta3}"].value:
                    sheet_ranges_smeta.merge_cells(f'A{start_index_table_templ}:I{start_index_table_templ}')
                    sheet_ranges_smeta[f'A{start_index_table_templ}'].alignment = Alignment(horizontal='left')
                    flag = False

                if flag:
                    sheet_ranges_smeta[f"A{start_index_table_templ}"].font = Font(name='Arial', size=9, bold=False)
                    sheet_ranges_smeta[f'A{start_index_table_templ}'].alignment = Alignment(horizontal='left')
                    sheet_ranges_smeta.merge_cells(f'A{start_index_table_templ}:M{start_index_table_templ}')

                if loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].font.bold:
                    sheet_ranges_smeta[f"A{start_index_table_templ}"].font = Font(name='Arial', bold=True)
                    sheet_ranges_smeta[f"J{start_index_table_templ}"].font = Font(name='Arial', bold=True)
                    if ('Итог' in loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value
                        or 'ВСЕГО' in loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value
                        or 'Всего' in loaded_sheet_ranges_smeta[f"A{start_index_smeta3}"].value):
                        sheet_ranges_smeta[f"A{start_index_table_templ}"].font = Font(name='Arial', size=9, bold=True)
                        sheet_ranges_smeta[f"J{start_index_table_templ}"].font = Font(name='Arial', size=8, bold=True)

                for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'):
                    sheet_ranges_smeta[f'{letter}{start_index_table_templ}'].border = Border(
                        top=black_side, left=black_side, right=black_side, bottom=black_side)

                start_index_table_templ += 1

        else:
            break


    for start_index_smeta4 in range(start_index_smeta3, 300):
        if loaded_sheet_ranges_smeta[f"A{start_index_smeta4}"].value:
            if 'Составил:' in loaded_sheet_ranges_smeta[f"A{start_index_smeta4}"].value:
                sheet_ranges_smeta.merge_cells(f'A{start_index_table_templ}:M{start_index_table_templ}')
                sheet_ranges_smeta[f'A{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"A{start_index_smeta4}"].value
                sheet_ranges_smeta[f"A{start_index_table_templ}"].font = Font(name='Arial', size=9, bold=False)
                sheet_ranges_smeta[f'A{start_index_table_templ}'].alignment = Alignment(horizontal='center')
                start_index_table_templ += 1
                continue

            if 'должность' in loaded_sheet_ranges_smeta[f"A{start_index_smeta4}"].value:
                sheet_ranges_smeta.merge_cells(f'A{start_index_table_templ}:M{start_index_table_templ}')
                sheet_ranges_smeta[f'A{start_index_table_templ}'].value = loaded_sheet_ranges_smeta[f"A{start_index_smeta4}"].value
                sheet_ranges_smeta[f"A{start_index_table_templ}"].font = Font(name='Arial', size=9, bold=False, italic=True)
                sheet_ranges_smeta[f'A{start_index_table_templ}'].alignment = Alignment(horizontal='center')
                start_index_table_templ += 1
                break

            start_index_table_templ += 1


        else:
            start_index_table_templ += 1

    f = BytesIO()
    wb.save(f)
    length = f.tell()
    f.seek(0)

    response = HttpResponse(
        f.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = "attachment; filename=KS2_KS3_smeta.xlsx"
    response['Content-Length'] = length

    return response
